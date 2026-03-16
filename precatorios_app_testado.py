import re
import unicodedata
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

st.set_page_config(page_title="Conversor de Precatórios", layout="wide")
st.title("Conversor de Planilha de Precatórios")
st.write("Envie a planilha .xlsx baixada do sistema e baixe o arquivo final limpo.")


COLUNAS_FINAIS = [
    "ORDEM DE PAGAMENTO",
    "MOMENTO DE APRESENTAÇÃO DO PRECATÓRIO",
    "PROCESSO",
    "PRECATÓRIO",
    "RP",
    "VENCIMENTO",
    "EXEQUENTE",
    "CPF",
    "VALOR DEVIDO / SALDO A PAGAR POR EXEQUENTE",
    "TIPO DE PREFERÊNCIA",
]


def limpar(valor):
    if valor is None or pd.isna(valor):
        return ""
    texto = str(valor).replace("\n", " ").replace("\r", " ").strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def normalizar_texto(texto):
    texto = limpar(texto).upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def separar_exequente_cpf(texto):
    texto = limpar(texto)
    if not texto:
        return "", ""

    # caso padrão: NOME - 000.000.000-00
    m = re.search(r"^(.*?)\s*-\s*(\d{3}\.\d{3}\.\d{3}-\d{2})", texto)
    if m:
        nome = limpar(m.group(1))
        cpf = m.group(2)
        return nome, cpf

    # tenta achar CPF em qualquer parte do texto
    m2 = re.search(r"(\d{3}\.\d{3}\.\d{3}-\d{2})", texto)
    if m2:
        cpf = m2.group(1)
        nome = limpar(texto.replace(cpf, "").replace(" - ", " ").replace("-", " "))
        return nome, cpf

    partes = re.split(r"\s*-\s*", texto, maxsplit=1)
    if len(partes) == 2:
        return limpar(partes[0]), limpar(partes[1])

    return texto, ""


def normalizar_valor_monetario(valor):
    texto = limpar(valor)

    if not texto:
        return None

    texto = texto.replace("R$", "").replace(" ", "")
    texto = texto.replace(".", "").replace(",", ".")

    try:
        return float(texto)
    except ValueError:
        return None


def carregar_linhas(arquivo):
    wb = load_workbook(arquivo, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def encontrar_linha_cabecalho(linhas):
    """
    Procura a linha que contém o cabeçalho principal da tabela.
    """
    for i, linha in enumerate(linhas):
        vals = [normalizar_texto(v) for v in linha]

        if (
            "ORDEM DE PAGAMENTO" in vals
            and "PROCESSO" in vals
            and any("TIPO DE PREFERENCIA" in v for v in vals)
        ):
            return i

    return None


def construir_mapa_colunas(linha_cabecalho):
    """
    Cria um dicionário {NOME_NORMALIZADO_DO_CABECALHO: índice_da_coluna}
    """
    mapa = {}
    for idx, valor in enumerate(linha_cabecalho):
        chave = normalizar_texto(valor)
        if chave:
            mapa[chave] = idx
    return mapa


def localizar_coluna(mapa, nomes_exatos=None, contem=None, obrigatoria=True):
    nomes_exatos = nomes_exatos or []
    contem = contem or []

    # tenta por nome exato
    for nome in nomes_exatos:
        nome_norm = normalizar_texto(nome)
        if nome_norm in mapa:
            return mapa[nome_norm]

    # tenta por "contém"
    for chave, idx in mapa.items():
        for trecho in contem:
            trecho_norm = normalizar_texto(trecho)
            if trecho_norm in chave:
                return idx

    if obrigatoria:
        raise ValueError(
            f"Não foi possível localizar a coluna. Procurado por: "
            f"{nomes_exatos if nomes_exatos else contem}"
        )
    return None


def obter_valor(vals, idx):
    if idx is None:
        return ""
    if idx < len(vals):
        return vals[idx]
    return ""


def converter_planilha(arquivo):
    linhas = carregar_linhas(arquivo)

    idx_header = encontrar_linha_cabecalho(linhas)
    if idx_header is None:
        raise ValueError(
            "Não foi possível localizar o cabeçalho da planilha. "
            "Verifique se o layout do arquivo é o esperado."
        )

    linha_cabecalho = linhas[idx_header]
    mapa = construir_mapa_colunas(linha_cabecalho)

    idx_ordem = localizar_coluna(
        mapa,
        nomes_exatos=["ORDEM DE PAGAMENTO"]
    )
    idx_momento = localizar_coluna(
        mapa,
        nomes_exatos=["MOMENTO DE APRESENTAÇÃO DO PRECATÓRIO"]
    )
    idx_processo = localizar_coluna(
        mapa,
        nomes_exatos=["PROCESSO"]
    )
    idx_precatorio = localizar_coluna(
        mapa,
        nomes_exatos=["PRECATÓRIO", "PRECATORIO"]
    )
    idx_rp = localizar_coluna(
        mapa,
        nomes_exatos=["RP"]
    )
    idx_vencimento = localizar_coluna(
        mapa,
        nomes_exatos=["VENCIMENTO"]
    )

    idx_exequente_fonte = localizar_coluna(
        mapa,
        nomes_exatos=["EXEQUENTE"],
        contem=["EXEQUENTE"],
        obrigatoria=True
    )

    idx_valor = localizar_coluna(
        mapa,
        nomes_exatos=["VALOR DEVIDO / SALDO A PAGAR POR EXEQUENTE"],
        contem=["SALDO A PAGAR POR EXEQUENTE", "VALOR DEVIDO"],
        obrigatoria=True
    )

    idx_tipo_pref = localizar_coluna(
        mapa,
        nomes_exatos=["TIPO DE PREFERÊNCIA", "TIPO DE PREFERENCIA"],
        contem=["TIPO DE PREFERENCIA"],
        obrigatoria=True
    )

    registros = []

    # guarda os últimos valores válidos das colunas que costumam vir mescladas
    ultimos_compartilhados = {
        "ordem": "",
        "momento": "",
        "processo": "",
        "precatorio": "",
        "rp": "",
        "vencimento": "",
        "tipo_pref": "",
    }

    for i, linha in enumerate(linhas[idx_header + 1:], start=idx_header + 1):
        vals = [limpar(v) for v in linha]

        if not any(vals):
            continue

        texto_linha = " | ".join([v for v in vals if v])
        texto_upper = normalizar_texto(texto_linha)

        ordem_bruta = limpar(obter_valor(vals, idx_ordem))

        # ignora cabeçalhos repetidos ou linhas institucionais
        if (
            ordem_bruta == "ORDEM DE PAGAMENTO"
            or "LISTA CONSOLIDADA - OFICIOS PRECATORIOS" in texto_upper
            or texto_upper.startswith("MUNICIPIO DE")
            or "PODER JUDICIARIO" in texto_upper
            or "TRIBUNAL REGIONAL" in texto_upper
            or "SECRETARIA DE PRECATORIOS" in texto_upper
        ):
            continue

        # lê os valores da linha
        momento = limpar(obter_valor(vals, idx_momento))
        processo = limpar(obter_valor(vals, idx_processo))
        precatorio = limpar(obter_valor(vals, idx_precatorio))
        rp = limpar(obter_valor(vals, idx_rp))
        vencimento = limpar(obter_valor(vals, idx_vencimento))
        tipo_preferencia = limpar(obter_valor(vals, idx_tipo_pref))
        exequente_fonte = limpar(obter_valor(vals, idx_exequente_fonte))
        valor_bruto = obter_valor(vals, idx_valor)

        # preenche campos compartilhados quando vierem vazios por causa de célula mesclada
        if ordem_bruta:
            ultimos_compartilhados["ordem"] = ordem_bruta
        else:
            ordem_bruta = ultimos_compartilhados["ordem"]

        if momento:
            ultimos_compartilhados["momento"] = momento
        else:
            momento = ultimos_compartilhados["momento"]

        if processo:
            ultimos_compartilhados["processo"] = processo
        else:
            processo = ultimos_compartilhados["processo"]

        if precatorio:
            ultimos_compartilhados["precatorio"] = precatorio
        else:
            precatorio = ultimos_compartilhados["precatorio"]

        if rp:
            ultimos_compartilhados["rp"] = rp
        else:
            rp = ultimos_compartilhados["rp"]

        if vencimento:
            ultimos_compartilhados["vencimento"] = vencimento
        else:
            vencimento = ultimos_compartilhados["vencimento"]

        if tipo_preferencia:
            ultimos_compartilhados["tipo_pref"] = tipo_preferencia
        else:
            tipo_preferencia = ultimos_compartilhados["tipo_pref"]

        # só considera linhas que tenham ordem válida após o preenchimento
        if not re.fullmatch(r"\d+", ordem_bruta):
            continue

        # só considera linhas que tenham um exequente
        if not exequente_fonte:
            continue

        nome, cpf = separar_exequente_cpf(exequente_fonte)
        valor_pago = normalizar_valor_monetario(valor_bruto)

        registros.append({
            "_linha_original": i,
            "ORDEM DE PAGAMENTO": ordem_bruta,
            "MOMENTO DE APRESENTAÇÃO DO PRECATÓRIO": momento,
            "PROCESSO": processo,
            "PRECATÓRIO": precatorio,
            "RP": rp,
            "VENCIMENTO": vencimento,
            "EXEQUENTE": nome,
            "CPF": cpf,
            "VALOR DEVIDO / SALDO A PAGAR POR EXEQUENTE": valor_pago,
            "TIPO DE PREFERÊNCIA": tipo_preferencia,
        })

    df_final = pd.DataFrame(registros)

    if df_final.empty:
        raise ValueError(
            "Nenhum registro foi extraído. Verifique se o layout da planilha é o mesmo do arquivo esperado."
        )

    # preserva a ordem original da planilha
    df_final = df_final.sort_values("_linha_original", kind="mergesort")
    df_final = df_final.drop(columns="_linha_original")

    # garante a ordem final das colunas
    df_final = df_final[COLUNAS_FINAIS]

    return df_final


arquivo = st.file_uploader("Upload da planilha .xlsx", type=["xlsx"])

if arquivo is not None:
    try:
        df_final = converter_planilha(arquivo)

        st.success(f"Planilha processada com sucesso. Registros extraídos: {len(df_final)}")
        st.dataframe(df_final, use_container_width=True)

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df_final.to_excel(writer, index=False, sheet_name="Final")

            ws = writer.sheets["Final"]

            # aplica formato numérico na coluna de valor, sem símbolo R$
            col_valor = COLUNAS_FINAIS.index("VALOR DEVIDO / SALDO A PAGAR POR EXEQUENTE") + 1
            for row in range(2, len(df_final) + 2):
                ws.cell(row=row, column=col_valor).number_format = '#,##0.00'

        st.download_button(
            label="Baixar planilha final",
            data=output.getvalue(),
            file_name="2_Final.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        st.error(f"Erro ao processar a planilha: {e}")